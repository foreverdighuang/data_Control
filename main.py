import openpyxl
import re
import os
from pathlib import Path
from openpyxl.styles import PatternFill
import time

# excle表格文件列常量定义
NAME=3 #文件名所在列
DIR=11 #文件路径所在列
DATE=6 #入库时间所在列

WORK_DIR=r"E:\积累" #文件库的位置
FILE_DIR=r"C:\Users\huang\Desktop\日常与深入学习\积累资料库文件清单.xlsx" #生成的清单文件位置
# 现使用的终端编号，笔记本7，家8，单位内网9，单位外网10   
CURRENT_PC=7



wb1=openpyxl.load_workbook(FILE_DIR) 
ws1=wb1.get_sheet_by_name(wb1.get_sheet_names()[0])#只有一个工作表

# 定义元素集合，提高运行效率
jihe=set(range(2,ws1.max_row+1))
currnet_Rows=ws1.max_row

# 定义递归函数，遍历所有子目录，并将文件写入表格
def find_file(dir_Name):
    for files  in os.listdir(dir_Name):
        if os.path.isdir(os.path.join(dir_Name,files)):
            if files[0:4] == "sub_":# 如果文件夹以sub_开头，则存储文件夹名1
                write_X(files,dir_Name)
                continue
            find_file(os.path.join(dir_Name,files))
        else:
            write_X(files,dir_Name)


# 定义写入表格函数
def write_X(name,dir):
    dir=dir.replace(WORK_DIR,".")#后续有需要生成决定路径，替换.\即可
    for r in jihe:
        if name == ws1.cell(row=r,column=NAME).value:
            if dir == ws1.cell(row=r,column=DIR).value:
                jihe.remove(r)#集合中删除重复的内容，效率提升为n！
                ws1.cell(row=r,column=CURRENT_PC).value="是"
                return
        r+=1
    ws1.cell(row=ws1.max_row+1,column=NAME).value=name
    ws1.cell(row=ws1.max_row,column=DIR).value=dir
    #ws1.cell(row=ws1.max_row,column=DIR).value=dir.replace(WORK_DIR,".")
    ws1.cell(row=ws1.max_row,column=DATE).value=time.strftime('%Y-%m-%d',time.localtime(time.time()))
    ws1.cell(row=ws1.max_row,column=CURRENT_PC).value="是"
    if len(str(dir).split('\\'))>1:
        ws1.cell(row=ws1.max_row,column=1).value=str(dir).split('\\')[1] # 输入类别1
    if len(str(dir).split('\\'))>2:
        ws1.cell(row=ws1.max_row,column=2).value=str(dir).split('\\')[2] # 输入类别2


find_file(WORK_DIR)
# 将没有查找到的文件对应行单元格标黄
for r in jihe:
    ws1.cell(row=r,column=NAME).fill=PatternFill(fill_type='solid', fgColor="FFC125")
    ws1.cell(row=r,column=CURRENT_PC).value="否"
wb1.save(FILE_DIR)







